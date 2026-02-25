import json
from openpyxl import load_workbook
import re

def audyt_excela(filepath):
    wb = load_workbook(filepath)
    raport_bledow = []

    for sheet in wb.sheetnames:
        obecny_arkusz = wb[sheet]
        if obecny_arkusz.sheet_state != "visible":
            raport_bledow.append(dict([("typ_bledu", "ukryty_arkusz")]))
        
        for row in obecny_arkusz.rows:
            for cell in row:
                komorka = cell.value
                lokalizacja = cell.coordinate
                if komorka is not None and str(komorka).startswith("="):
                    znalezione_hardkodowanie = re.findall(r'\b\d+(\.\d+)?', komorka)
                    if znalezione_hardkodowanie:
                        raport_bledow.append(dict([("typ_bledu", "hardkodowanie"), ("komorka", lokalizacja)]))
                
                if isinstance(komorka, str) and komorka != komorka.strip():
                    raport_bledow.append(
                        {"typ_bledu": "zbedne spacje", "komorka": lokalizacja}
                    )
        return json.dumps(raport_bledow, indent=4, ensure_ascii=False)